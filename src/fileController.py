# fileController, [file] - file значит не само понятие файла, 
#                           а операцию обработки файла в целом

from collections import namedtuple
import datetime
from typing import List
from openpyxl.worksheet.worksheet import Worksheet
from time import sleep
import traceback
import re

from excelDriver import ExcelDriver
from label import LabelController
from learning import LearningDriver, UserNotFound
from datatypes import UserAction, UserActionType, UserInfo, AuthCookies
from utils import (
    generate_random_string, suggest_user_actions, 
    convert_date_string, is_blue_color, is_red_color
)
from settings import Settings

    
class FileController:
    @staticmethod
    def step1(filepath, progress_gen, ask_user_actions, confirm_users_actions, message_callback):
        """ Обработка файла часть 1
            Args:
                filepath (str): Путь к excel файлу
                progress_gen: Callback-функция, принимающая итерируемый объект
                                и возвращающая генератор.
                                В интерфейсе показывает прогресс-бар
                ask_user_actions (Callable): Callback для выбора действий над пользователем
                confirm_user_actions (Callable): Callback для подтверждения действий
                message_callback (Callable): Callback для отправки сообщений
        """
        learning = LearningDriver(AuthCookies(*Settings().get_crypted('auth')))
        xlsx = ExcelDriver()
        xlsx.load(filepath)
        
        # Очистка лишних листов
        ws_general = xlsx.get_first_worksheet()
        xlsx.remove_other_sheets(ws_general)
        xlsx.save()
        
        # Перезагрузка
        xlsx.load(filepath)
        ws_general = xlsx.get_first_worksheet()

        # Лист 'Общий' или 'Лист1'
        ws_general.title = 'Общий'
        xlsx.insert_passwords(ws_general)

        # Лист 'Для предметов и меток'
        ws_labels = xlsx.clone_sheet(ws_general)
        ws_labels.title = 'Для предметов и меток'

        # Обработка пользователей
        user_table_data = xlsx.get_all_users_data()
        users_exists = list() # все зарегистрированные пользователи
        for table_user in progress_gen(user_table_data, title="Поиск пользователей..."):
            sleep(0.05)

            try:
                # list, тк по одному email может быть несколько пользователей
                uinfo_list = learning.get_user_info(table_user.email)
            except UserNotFound:
                continue

            for uinfo in uinfo_list:
                uinfo.table = table_user
                users_exists.append(uinfo)
        
        # Отправка сообщения о завершении загрузки пользователей
        _t1 = len(user_table_data)
        _t2 = len(users_exists)
        _t3 = round(_t2 / _t1 * 100, 2)
        message_callback(f"Найдено {_t2}/{_t1} ({_t3}%) пользователей", status="info")
        
        if not _t2:
            message_callback(f"Зарегистрированных пользователей нет.", status="info")
        else:
            message_callback(f"Выберите действия для найденных пользователей...")
            sleep(1.5)

            # Выбор судьбы пользователей
            user_actions = list() # список действий над пользователями
            for userinfo in users_exists:
                suggested = suggest_user_actions(userinfo, learning=learning)
                uactions = ask_user_actions(userinfo, suggested)
                user_actions.append((userinfo, uactions))

            # Подтверждение
            users_actions_confirmed = confirm_users_actions(user_actions)
            if not users_actions_confirmed:
                message_callback("Обработка прервана.", status='bad')
                return

            # Реализация судьбы пользователей
            for user_action in progress_gen(user_actions, title="Выполнение действий..."):
                try:
                    FileController.perform_user_actions(
                        xlsx, 
                        learning, user_action[0], user_action[1]
                    )
                except Exception as e:
                    message_callback(traceback.format_exc(), status="info")
                    message_callback(f"Не удалось выполнить действия для пользователя ({user_action[0].mid}, {user_action[0].email})", status="bad")
                    message_callback(user_action[1], status='info')
            
        # Финиш
        ws_logins = xlsx.create_sheet(title="Для логинов", index=1)
        xlsx.clone_sheet_unique(ws_copy=ws_labels, ws_paste=ws_logins, unique_column_name='email')
        xlsx.save()
        message_callback("Обработка пользователей завершена. Файл сохранён.")

    @staticmethod
    def progresstest(callback):
        import time
        for i in callback([x for x in range(25)]):
            print("Doing action...", i)
            time.sleep(0.5)

    @staticmethod
    def completetest(callback):
        callback("test message")

    @staticmethod
    def perform_user_actions(xlsx: ExcelDriver, 
            learning: LearningDriver, uinfo: UserInfo, uacts: List[UserAction]):
        for uact in uacts:
            if uact.completed: continue
            xlsx_changed = False
            if uact == UserActionType.SKIP:
                xlsx.mark_user_as_skipped(uinfo.table.email)
                xlsx_changed = True
            elif uact == UserActionType.DELETE:
                learning.delete(uinfo.mid)
            elif uact == UserActionType.DELETE_FROM_TABLE:
                xlsx.delete_user_from_workbook(uinfo.table.email)
            elif uact == UserActionType.ADD_LABEL:
                learning.add_tag(uinfo.mid, uact.param)
            elif uact == UserActionType.REMOVE_LABEL:
                learning.remove_tag(uinfo.mid, uact.param)
            elif uact == UserActionType.CHANGE_LOGIN:
                login = uinfo.login
                # try:
                #     password = (int(uinfo.table.login[-5:])+23000)*15
                # except (ValueError, AttributeError):
                #     password = generate_random_string(8)
                # learning.set_password(uinfo.mid, password)
                # xlsx.change_login_password(uinfo.table.email, login, password)
                xlsx.change_columns(uinfo.table.email, [('логин', login)])
                xlsx.mark_user_as_skipped(uinfo.table.email)
                xlsx_changed = True
            elif uact == UserActionType.CHANGE_PASSW_EDU:
                learning.set_password(uinfo.mid, uact.param)
            elif uact == UserActionType.CHANGE_PASSW_LOCAL:
                xlsx.change_columns(uinfo.table.email, [('пароль', uact.param)])
                xlsx_changed = True
            elif uact == UserActionType.MARK_REGISTERED:
                xlsx.mark_user_as_registered(uinfo.table.email)
                xlsx_changed = True
            elif uact == UserActionType.SET_COMMENT:
                xlsx.set_comment(xlsx.get_first_worksheet(), uinfo.table.email, uact.param)
                xlsx_changed = True
            if xlsx_changed:
                xlsx.save()
            uact.completed = True

    @staticmethod
    def step2(filepath, message_callback):
        driver = ExcelDriver()
        driver.load(filepath)
        
        if not 'Для предметов и меток' in driver._xlsx.sheetnames:
            message_callback("Отсутствует лист Для предметов и меток", status='bad')
            return False
        
        ws_labels = driver._xlsx['Для предметов и меток']
        ws_labels_users = {}
        cols = {
            'email': driver.get_column_by_name(ws_labels, 'email') - 1,
            'surname': driver.get_column_by_name(ws_labels, 'ФИО') - 1,
            'name': driver.get_column_by_name(ws_labels, 'ФИО') + 1 - 1,
            'patronymic': driver.get_column_by_name(ws_labels, 'ФИО') + 2 - 1,
            'login': driver.get_column_by_name(ws_labels, 'Логин') - 1,
            'password': driver.get_column_by_name(ws_labels, 'Пароль') - 1,
            'admission_code': driver.get_column_by_name(ws_labels, 'Код поступ.') - 1,
            'subject_name': driver.get_column_by_name(ws_labels, 'Предмет') - 1,
            'subject_date': driver.get_column_by_name(ws_labels, 'Выбранная дата') - 1,
        }
        User = namedtuple('User', [
            'email', 'surname', 'name', 'patronymic', 
            'login', 'password', 'admission_code', 'subjects'
        ])

        for row in ws_labels.iter_rows(min_row=2):
            fill = row[cols['surname']].fill.fgColor
            if fill.type == 'theme':
                if fill.value not in [9,7,6]:
                    raise Exception('unknown theme', fill)
            elif fill.type == 'rgb':
                if is_blue_color(fill.value): continue
                if is_red_color(fill.value): continue
            email = row[cols['email']].value
            user = ws_labels_users.get(email)
            if not user:
                user = User(
                    email=email, 
                    surname=row[cols['surname']].value, name=row[cols['name']].value, 
                    patronymic=row[cols['patronymic']].value,
                    login=row[cols['login']].value, password=row[cols['password']].value,
                    admission_code=row[cols['admission_code']].value, subjects=[]
                )
                ws_labels_users[email] = user
            user.subjects.append((row[cols['subject_name']].value, row[cols['subject_date']].value))
        
        if '_csv' in driver._xlsx.sheetnames:
            driver._xlsx.remove(driver._xlsx['_csv'])
            driver.save()
            driver.load(filepath)

        ws = driver.create_sheet(title="_csv")
        header = "Табельный номер;Фамилия;Имя;Отчество;Login;E-mail;Пароль;Преподаватель;Группа;Метки".split(';')
        ws.append(header)

        reexp = re.compile(r'=\(RIGHT\(..+(;|,)5\)\+23000\)\*15')
        for user in ws_labels_users.values():
            labels = []
            for subject_tuple in user.subjects:
                subject = subject_tuple[0]
                date = convert_date_string(str(subject_tuple[1])).date() if subject_tuple[1] else None
                labels.append(LabelController.get_label(subject, selected_date=date))
            labels = ','.join(labels)
            password = (int(user.admission_code[-5:])+23000)*15 if reexp.match(user.password) else user.password
            ws.append((
                user.admission_code, user.surname, user.name, user.patronymic,
                user.login, user.email, password, 0, '', labels
            ))

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['J'].width = 48
        
        driver.save()
        message_callback("Обработка файла завершена. Файл сохранен.")



            

