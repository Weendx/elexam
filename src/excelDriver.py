from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import os
import re
import errno
from typing import List, Tuple, NamedTuple, Optional, Iterable
from copy import copy

from datatypes import EmailNLogin, UserTableData, TableSubject
from utils import convert_date_string, is_array_consecutive, is_blue_color, is_red_color


COLOR_FILL_REGISTERED = 'FF558ED5'
COLOR_FILL_SKIPPED = 'FFFF3838'
COLOR_FILL_COMMENT = 'FFA9D08E'

class NotLoadedException(Exception):
    pass

class ColumnNotFoundException(Exception):
    pass

class UserNotFoundException(Exception):
    pass

class SheetNotFoundException(Exception):
    pass

class ExcelDriver:

    _xlsx = None
    _filepath = None

    def append_rows(self, rows: Iterable[Iterable], worksheet=None) -> bool:
        self.check_loaded()
        ws = worksheet if worksheet else self._xlsx.active

        rows_count = len(rows)
        if not rows_count: return False
        cols_count = len(rows[0])
        if not cols_count: return False
        
        for row in rows:
            ws.append(row)
        
        return True

    @classmethod
    def apply_row_style(cls, worksheet, row, style):
        """ Update style of the first 9th cells in the row """
        
        styles = ["fill", "font", "border", "number_format", "protection", "alignment"]
        for cell_range in worksheet.iter_rows(min_row=row, max_row=row, max_col=9):
            for cell in cell_range:
                for style_tag in styles:
                    if style_tag in style:
                        setattr(cell, style_tag, style[style_tag])

    def change_columns(self, email: str, changes: List[Tuple[str, str]]):
        """ Change values in columns
            changes: [(column_name, value), ...] 
        """

        self.check_loaded()

        for ws in self._xlsx:
            email_col = self.get_email_column_id(ws)
            column_numbers = dict()

            for change in changes:
                try:
                    column_numbers[change[0]] = self.get_column_by_name(ws, change[0])
                except ColumnNotFoundException:
                    continue

            for row in ws.iter_rows():
                email_cell = row[email_col - 1]            
                if email_cell.value != email: continue
                for change in changes:
                    col = column_numbers.get(change[0])
                    if not col: continue
                    cell = row[col - 1]
                    cell.value = change[1]

    def change_login_password(self, email: str, login: str, password: str):
        self.check_loaded()
        for sheetname in self._xlsx.sheetnames:
            ws = self._xlsx[sheetname]

            try:
                email_col = self.get_email_column_id(ws)
                login_col = self.get_column_by_name(ws, 'логин')
                passw_col = self.get_column_by_name(ws, 'пароль')
            except ColumnNotFoundException:
                continue

            max_col = max((email_col, login_col, passw_col))
            for row in ws.iter_rows(max_col=max_col):
                email_cell = row[email_col - 1]
                login_cell = row[login_col - 1]
                passw_cell = row[passw_col - 1]                
                if email_cell.value != email: continue
                login_cell.value = login
                passw_cell.value = password

    def check_loaded(self):
        if self._xlsx == None:
            raise NotLoadedException()

    def clone_sheet(self, source_worksheet) -> Worksheet:
        """ Returns worksheet with cloned data (unique) """
        return self._xlsx.copy_worksheet(source_worksheet)

    def clone_sheet_unique(self, ws_copy, ws_paste, unique_column_name) -> None:
        """ Copy data from ws_copy to ws_paste (unique) """
        unique_column = self.get_column_by_name(ws_copy, unique_column_name)
        email_column = self.get_email_column_id(ws_copy)
        password_column = self.get_column_by_name(ws_copy, 'пароль')
        password_formula = self.get_password_formula()

        formula_reexp = re.compile(r'=\(RIGHT\(..+(;|,)5\)\+23000\)\*15')

        added = list()
        last_inserted_row = 1
        for i, row in enumerate(ws_copy.iter_rows()):
            if ws_copy.cell(row=i+1, column=email_column).value == '<deleted>': continue
            unique_cell = ws_copy.cell(row=i+1, column=unique_column)
            if unique_cell.value in added: continue
            for cell in row:
                new_cell = ws_paste.cell(row=last_inserted_row, column=cell.column,
                        value=cell.value)
                if cell.column == password_column and i != 0 and formula_reexp.match(cell.value):
                        new_cell.value = password_formula.replace('%i', str(last_inserted_row))
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
            ##
            last_inserted_row += 1
            added.append(unique_cell.value)

    def create_sheet(self, *args, **kwargs):
        self.check_loaded()
        return self._xlsx.create_sheet(*args, **kwargs)

    def delete_rows(self, worksheet: Worksheet, row: int, amount=1) -> None:
        """ Delete rows with formula translation for cells shifted up. """
        self.check_loaded()
        if amount < 1 or row > worksheet.max_row:
            return

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        amount = min(amount, max_row - row + 1)
        first_moved_row = row + amount
        if first_moved_row <= max_row:
            worksheet.move_range(
                f'A{first_moved_row}:{get_column_letter(max_column)}{max_row}',
                rows=-amount,
                translate=True,
            )
        worksheet.delete_rows(max_row - amount + 1, amount)

    def delete_row(self, worksheet: Worksheet, row: int) -> None:
        self.delete_rows(worksheet, row)

    def delete_user_from_workbook(self, email: str, subject: Optional[str] = None) -> None:
        self.check_loaded()

        general_worksheet = self._xlsx["Общий"]
        rows_in_general_list = self.get_rows_with_user(general_worksheet, email)

        for worksheet in self._xlsx:
            if not subject and worksheet == general_worksheet:
                rows = rows_in_general_list
            else:
                rows = self.get_rows_with_user(worksheet, email, subject=subject)
            
            if not rows:
                continue

            if subject and len(rows_in_general_list) != 1:
                if worksheet.title == "_csv":
                    # Помечаем красным чтобы потом проверить метки.
                    # По хорошему здесь добавить пересборку меток, но 
                    # для _csv нужно писать отдельный класс похоже.
                    self.mark_user(worksheet, email, True, COLOR_FILL_SKIPPED)
                    self.set_comment(worksheet, email, "Метки указаны некорректно") 
                    continue
                if worksheet.title == "Для логинов":
                    # Удаляем из логинов только если у человека один предмет чтобы не потерять
                    continue

            if is_array_consecutive(rows):
                self.delete_rows(worksheet, rows[0], len(rows))
            else:
                for row in rows:
                    self.delete_row(worksheet, row)
            
    def get_all_users_data(self, first_row_is_header=True) -> Tuple[UserTableData]:
        self.check_loaded()
        userdata = dict()
        ws = self.get_first_worksheet()

        email_column = self.get_email_column_id(ws)
        login_column = self.get_column_by_name(ws, 'логин')
        subject_column = self.get_column_by_name(ws, 'предмет')
        sel_date_column = self.get_column_by_name(ws, 'выбранная дата')
        fio_columns = self.get_columns_with_fio(ws)
        
        start_row = 2 if first_row_is_header else 1
        for column in ws.iter_cols(min_col=email_column, max_col=email_column, min_row=start_row):
            for cell in column:
                email = ws.cell(row=cell.row, column=email_column).value
                utd = userdata.get(email)
                if not utd:
                    fio = ""
                    for fio_subcolumn in fio_columns: 
                        value = ws.cell(row=cell.row, column=fio_subcolumn).value
                        if not value or not value.strip(): continue
                        if fio: value = ' ' + value
                        fio += value
                    if not fio: fio = None

                    mark = self.get_cell_mark(ws.cell(row=cell.row, column=fio_columns[0]))
                    marks = [mark] if mark != 'none' else None

                    userdata[email] = UserTableData(
                        email=email,
                        login=ws.cell(row=cell.row, column=login_column).value,
                        fio=fio,
                        subjects=[],
                        marks = marks
                    )
                    utd = userdata[email]
                subject_name = ws.cell(row=cell.row, column=subject_column).value
                subject_date = ws.cell(row=cell.row, column=sel_date_column).value
                subject_date = convert_date_string(str(subject_date)) if subject_date else None
                utd.subjects.append(TableSubject(subject_name, subject_date))

        return tuple(userdata.values())

    @classmethod
    def get_cell_mark(cls, cell) -> str:
        """ Получить кодовое обозначение цветовой отметки ячейки
            Returns: none | registered | skipped | commented
        """
        fill = cell.fill.fgColor
        if fill.type != 'rgb':
            # print Warning: unsupported fill type, use rgb fill
            return 'none'
        if is_blue_color(fill.value):
            return 'registered'
        if is_red_color(fill.value):
            return 'skipped'
        if fill.value == COLOR_FILL_COMMENT:
            return 'commented'
        return 'none'

    @classmethod
    def get_column_by_name(cls, worksheet, column_name) -> int:
        for row in worksheet.iter_rows(max_row=1):
            for cell in row:
                if not cell.value: continue
                if cell.value.lower().strip() == column_name.lower().strip():
                    return cell.column
        raise ColumnNotFoundException()

    @classmethod
    def get_email_column_id(cls, worksheet) -> int:
        try:
            return cls.get_column_by_name(worksheet, 'email')
        except ColumnNotFoundException:
            return cls.get_column_by_name(worksheet, 'e-mail')
        
    @classmethod
    def get_columns_with_fio(cls, worksheet) -> Tuple[int]:
        """ Returns a tuple with fio columns
            Example: (2,3,4)
        """
        fio_last_name = cls.get_column_by_name(worksheet, 'фио')
        return (fio_last_name, fio_last_name+1, fio_last_name+2)

    @classmethod
    def get_rows_with_user(cls, worksheet, email: str, subject: Optional[str] = None) -> Tuple[int]:
        """ Returns a tuple with rows id containing the specified user with email.
            Can be filtered by subject.

            Example: (1,2,3)
        """
        rows_list = []
        start_row = 1
        email_column = cls.get_email_column_id(worksheet)
        email = str(email).lower().strip()
        subject = str(subject).lower().strip() if subject else None
        subject_column = None

        try:
            subject_column = cls.get_column_by_name(worksheet, "Предмет") if subject else subject_column
        except ColumnNotFoundException:
            pass

        for row in worksheet.iter_cols(min_col=email_column, max_col=email_column, min_row=start_row):
            for cell in row:
                if str(cell.value).lower().strip() != email:
                    continue
                if subject_column:
                    subject_cell = worksheet.cell(row=cell.row, column=subject_column)
                    if str(subject_cell.value).lower().strip() != subject:
                        continue
                rows_list.append(cell.row)
        return rows_list


    def get_emails(self, worksheet, first_row_is_header=True):
        email_column = self.get_email_column_id(worksheet)
        start_row = 2 if first_row_is_header else 1
        email_generator = worksheet.iter_cols(min_row=start_row, min_col=email_column, 
                                                max_col=email_column, values_only=True)
        return set(next(email_generator))
    
    def get_emails_n_logins(self, worksheet, first_row_is_header=True):
        email_column = self.get_email_column_id(worksheet)
        login_column = self.get_column_by_name(worksheet, 'логин')
        max_col = email_column if email_column > login_column else login_column
        start_row = 2 if first_row_is_header else 1
        generator = worksheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True)
        output = set()
        for row in generator:
            output.add(EmailNLogin(email=row[email_column - 1], login=row[login_column - 1]))
        return output
    
    def get_first_worksheet(self):
        self.check_loaded()
        ws = self._xlsx.active
        
        firstlist_names = ['лист 1', 'лист1', 'общий', 'sheet 1', 'sheet1']
        if ws.title.lower() in firstlist_names:
            return ws

        for sheetname in self._xlsx.sheetnames:
            if sheetname.lower() in firstlist_names:
                return self._xlsx[sheetname]
        raise SheetNotFoundException

    @staticmethod
    def get_password_formula() -> str:
        """ Returns a formula for password generation 
            
            Note: 
                use , as argument separator instead of ; in formula
                  %i for row number
            Example:
                =(RIGHT(H%i,5)+23000)*15
        """
        return "=(RIGHT(H%i,5)+23000)*15"

    def get_user_data(self, email, first_row_is_header=True) -> UserTableData:
        self.check_loaded()
        _email = email
        userdata = None
        ws = self.get_first_worksheet()

        email_column = self.get_email_column_id(ws)
        login_column = self.get_column_by_name(ws, 'логин')
        subject_column = self.get_column_by_name(ws, 'предмет')
        sel_date_column = self.get_column_by_name(ws, 'выбранная дата')
        fio_columns = self.get_columns_with_fio(ws)
        
        start_row = 2 if first_row_is_header else 1
        for column in ws.iter_cols(min_col=email_column, max_col=email_column, min_row=start_row):
            for cell in column:
                email = ws.cell(row=cell.row, column=email_column).value
                if email != _email: continue
                if not userdata:
                    fio = ""
                    for fio_subcolumn in fio_columns: 
                        value = ws.cell(row=cell.row, column=fio_subcolumn).value
                        if not value or not value.strip(): continue
                        if fio: value = ' ' + value
                        fio += value
                    if not fio: fio = None

                    mark = self.get_cell_mark(ws.cell(row=cell.row, column=fio_columns[0]))
                    marks = [mark] if mark != 'none' else None

                    userdata = UserTableData(
                        email=email,
                        login=ws.cell(row=cell.row, column=login_column).value,
                        fio=fio,
                        subjects=[],
                        marks=marks
                    )
                subject_name = ws.cell(row=cell.row, column=subject_column).value
                subject_date = ws.cell(row=cell.row, column=sel_date_column).value
                subject_date = convert_date_string(str(subject_date)) if subject_date else None
                userdata.subjects.append(TableSubject(subject_name, subject_date))

        return userdata

    def get_worksheet(self, name):
        self.check_loaded()
        for sheetname in self._xlsx.sheetnames:
            if sheetname.lower() == name.lower():
                return self._xlsx[sheetname]
        return None

    def insert_passwords(self, worksheet, first_row_is_header=True):
        """ Insert password formula into the sheet for password column 
        
            Note: use , as argument separator instead of ; in formula
                  %i for row number
            formula = ExcelDriver.get_password_formula()
        """
        formula = self.get_password_formula()
        password_column_id = self.get_column_by_name(worksheet, 'пароль')
        start_row = 2 if first_row_is_header else 1
        password_column = next(worksheet.iter_cols(min_row=start_row, 
                                min_col=password_column_id, max_col=password_column_id))
        for cell in password_column:
            cell.value = formula.replace('%i', str(cell.row))

    def create_empty(self):
        self._xlsx = Workbook()

    def load(self, filepath):
        if not os.path.isfile(filepath):
            raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), filepath)
        self._xlsx = load_workbook(filepath)
        self._filepath = filepath

    def remove_other_sheets(self, worksheet):
        self.check_loaded()
        for sheet in self._xlsx:
            if sheet != worksheet:
                self._xlsx.remove(sheet)

    def save(self, filepath=None):
        self.check_loaded()
        _filepath = filepath if filepath else self._filepath
        self._xlsx.save(_filepath)
        self._filepath = _filepath

    @classmethod
    def mark_user(cls, worksheet, email, first_row_is_header=True, fgColor='FF558ED5'):
        """ Mark users who should not be registered by blue color """
        fill = PatternFill('solid', fgColor=fgColor.upper())
        start_row = 2 if first_row_is_header else 1
        email_column = cls.get_email_column_id(worksheet)

        for row in worksheet.iter_cols(min_col=email_column, max_col=email_column, min_row=start_row):
            for cell in row:
                if cell.value != email:
                    continue
                current_row = cell.row
                cls.apply_row_style(worksheet, current_row, {"fill": fill})

    def mark_user_as_registered(self, email, first_row_is_header=True):
        self.check_loaded()
        for worksheet in self._xlsx:
            self.mark_user(worksheet, email, first_row_is_header, COLOR_FILL_REGISTERED)
    
    def mark_user_as_skipped(self, email, first_row_is_header=True):
        self.check_loaded()
        for worksheet in self._xlsx:
            self.mark_user(worksheet, email, first_row_is_header, COLOR_FILL_SKIPPED)

    @classmethod
    def set_comment(cls, worksheet, email, comment) -> bool:
        """ Sets the comment at the first email cell. """
        fill = PatternFill('solid', fgColor=COLOR_FILL_COMMENT)
        email_column = cls.get_email_column_id(worksheet)
        for col_items in worksheet.iter_cols(min_col=email_column, max_col=email_column):
            for cell in col_items:
                if cell.value == email:
                    cell.comment = Comment(comment, 'elexam')
                    cell.fill = fill
                    return True
        return False

    def write_header(self, header: Iterable, cols_size: Optional[Iterable] = None, row_num=1, worksheet=None):
        self.check_loaded()
        ws = worksheet if worksheet else self._xlsx.active

        header_len = len(header)
        if cols_size and len(cols_size) != header_len:
            raise ValueError("Length of 'cols_size' should be equal to 'header' length")

        bold_font = Font(name='Calibri', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=row_num, max_row=row_num, max_col=header_len):
            for i, title in enumerate(header):
                row[i].value = title
                row[i].font = bold_font
                row[i].alignment = center_alignment
                if cols_size:
                    ws.column_dimensions[row[i].column_letter].width = cols_size[i]
